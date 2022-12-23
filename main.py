from datetime import date
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl
# driver = webdriver.Chrome()
# driver.get("https://www.kodlama.io/")
# driver.maximize_window() # -> ekranƒ± tam boyutuna getirir
## loginBtn texti "Giri≈ü Yap" olmalƒ±dƒ±r..

# sitenin ya da elementin y√ºklenmesi beklenmeli..
# WebDriverWait => condition bazlƒ± √ßalƒ±≈üƒ±r
# loginBtn g√∂r√ºn√ºr olana kadar maximum 5 saniye bekle..
# => condition maximum ka√ß saniye bekletilsin
# loginBtnFinder = (By.XPATH,"//*[@id='navbar']/div/div/div/ul/li[3]/a") #neye g√∂re locate olacaƒüƒ±m
# WebDriverWait(driver,5).until(expected_conditions.visibility_of_element_located(loginBtnFinder)) # defansif kodlama
# loginBtn = driver.find_element(loginBtnFinder)
# loginBtnText = loginBtn.text

# #windows + .
# if loginBtnText == "Giri≈ü Yap":
#     print("Test ba≈üarƒ±lƒ± üòé")
# else:
#     print("Test Ba≈üarƒ±sƒ±z ‚ùå")


# ekran g√∂r√ºnt√ºs√º alma
# kaydetme
# element bazlƒ±, ya da driver bazlƒ± alƒ±nabilir.
# loginBtn.screenshot("element.png") # => elementin screenshotunu alƒ±r
# driver.save_screenshot(str(date.today()) + '(2).png') # => t√ºm ekranƒ±n ss'ini alƒ±r

# scroll_to fonksiyonu => locate edilmi≈ü bir element'e scroll yapar.
# javascript kullanƒ±larak scroll edilecek..  
# terms_conditions = driver.find_element(By.XPATH,'/html/body/div[1]/footer/div/div/div[2]/ul/li[1]/a')
#sleep(2)
#terms_conditions.screenshot(str(date.today()) + '(1).png')
# driver.execute_script("window.scroll(0,900)")
# driver.execute_script("arguments[0].scrollIntoView()",terms_conditions)

# ActionChains => Yapƒ±lacak aksiyonlarƒ± sƒ±rala, ve perform dediƒüinde bu i≈ülemler sƒ±rasƒ±yla
# ger√ßekle≈ütirilsin

# elemana scroll ol => screenshot al => butona tƒ±kla
# perform
#actions = ActionChains(driver)
#actions.move_to_element(terms_conditions)

# PG_DWN
# √∂zel karakterler nasƒ±l kullanƒ±lƒ±r => Caps,Shift,Ctrl,Alt,PG_DWN,PG_UP,Insert
#actions.send_keys(Keys.PAGE_DOWN)
#actions.click(terms_conditions)
#actions.perform() # => zincirlenen aksiyonlarƒ± i≈üleme koyar..


# 2 adet test case verilecek
# bu test caseler otomatize edilecek.
# sonu√ßlar console'a yazdƒ±rƒ±lacak
# ekran g√∂r√ºnt√ºs√º g√ºn√ºn tarihi ile kaydedilecek.
# date.today()

# WebDriverWait classƒ± üü©
# pytest => 
# pytest dosyalarƒ± "test_" prefixi (√∂n ek) ile ba≈ülar. 
# pytest classlarƒ± "Test_" prefixi ile ba≈ülar.
# pytest fonksiyonlar "test_" prefixi ile ba≈ülar.
# constants
excelFile = openpyxl.load_workbook("data/userFailData.xlsx")
selectedSheet = excelFile["Sheet1"]

rows = selectedSheet.max_row
columns = selectedSheet.max_column
      
for i in range(2,rows):
    username = selectedSheet.cell(i,1).value
    password = selectedSheet.cell(i,2).value
    print(username,password)
